import os
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import logging
import uuid
from datetime import datetime

def export_to_excel(data):
    """Export extracted data to Excel file in REFUND AUDIT LOG format."""
    try:
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "REFUND AUDIT LOG"
        
        # Add title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "REFUND AUDIT LOG SUMMARY"
        title_cell.font = openpyxl.styles.Font(bold=True, size=14)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Add headers - following the format in the image
        headers = ['Item #','Department', 'Qty', 'Total Sell', 'Period']
        
        # Apply header styling
        header_row = 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Add data rows - convert our extracted data to this format
        start_row = 3
        for row_num, item in enumerate(data, start_row):
            # Item number
            ws.cell(row=row_num, column=1).value = item.get('item_number', '')
            
            #Department
            ws.cell(row=row_num, column=2).value = item.get('department', '')
            
            # Quantity (use the quantity field if available, default to 1)
            # Quantity (handles AS400 '1-' format too)
            qty_raw = str(item.get('quantity', '1')).strip()

            if qty_raw.endswith('-') and qty_raw[:-1].isdigit():
                qty = int('-' + qty_raw[:-1])
            elif qty_raw.isdigit():
                qty = int(qty_raw)
            else:
                qty = 1

            ws.cell(row=row_num, column=3).value = qty

            
            # Total Sell (Price) - If we have quantity > 1, this should be the total amount
            price_cell = ws.cell(row=row_num, column=4)
            price_str = str(item.get('price', '0.00')).replace('Y', '').replace(',', '').strip()

            if price_str.endswith('-') and price_str[:-1].replace('.', '', 1).isdigit():
                price_val = -float(price_str[:-1])
            elif price_str.replace('.', '', 1).isdigit():
                price_val = float(price_str)
            else:
                price_val = 0.00

            price_cell.value = price_val
            price_cell.number_format = '$#,##0.00'


            
            # Period (use the period field if available, or extract from date)
            period = item.get('period', '')
            if not period:
                date_str = item.get('date', '')
                period = "P00"  # Default
                if date_str:
                    try:
                        # Try to extract month from date and convert to period
                        if '/' in date_str:
                            month = date_str.split('/')[0]
                            if month.isdigit():
                                period = f"P{month.zfill(2)}"
                    except Exception:
                        pass
            ws.cell(row=row_num, column=5).value = period
            

        
        # Add totals row if there is data
        if len(data) > 0:
            total_row = start_row + len(data)
            
            ws.cell(row=total_row, column=1).value = ""
            ws.cell(row=total_row, column=2).value = ""
            ws.cell(row=total_row, column=3).value = ""
            
            # Add "Grand Total" label
            ws.cell(row=total_row, column=4).value = "Grand Total"
            ws.cell(row=total_row, column=4).font = openpyxl.styles.Font(bold=True)
            
            # Sum the quantity columns
            qty_sum_formula = f"=SUM(C{start_row}:C{total_row-1})"
            ws.cell(row=total_row, column=5).value = f"=SUM(C{start_row}:C{total_row-1})"
            ws.cell(row=total_row, column=5).number_format = '0'

        
        # Auto-adjust column widths
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(i)
            for cell in col:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = max(max_length, 10) + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Make sure export directory exists
        export_dir = os.path.join('/tmp', 'exports')
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        
        # Save the workbook with a timestamp to avoid collisions
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"refund_audit_log_{timestamp}.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)
        
        logging.info(f"Data exported to Excel file: {filepath}")
        return filepath
    
    except Exception as e:
        logging.error(f"Error exporting to Excel: {str(e)}")
        raise e

def export_to_google_sheets(data):
    """Export extracted data to Google Sheets in REFUND AUDIT LOG format."""
    try:
        # Get Google Sheets credentials from environment or file
        credentials_json = os.environ.get('GOOGLE_CREDENTIALS')
        if not credentials_json:
            logging.warning("No Google credentials found in environment variables.")
            raise ValueError("Google Sheets credentials not found. Please check your configuration.")
        
        # Load credentials from JSON
        credentials_dict = json.loads(credentials_json)
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
        client = gspread.authorize(credentials)
        
        # Create a new spreadsheet
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        spreadsheet_title = f"REFUND AUDIT LOG - {timestamp}"
        spreadsheet = client.create(spreadsheet_title)
        
        # Get the first sheet and rename it
        worksheet = spreadsheet.get_worksheet(0)
        worksheet.update_title("REFUND AUDIT LOG")
        
        # Add title in merged cells (A1:G1)
        worksheet.merge_cells('A1:G1')
        worksheet.update_cell(1, 1, "REFUND AUDIT LOG SUMMARY")
        
        # Format title - need to use spreadsheet formatting features for this
        title_format = {
            "textFormat": {"bold": True, "fontSize": 14},
            "horizontalAlignment": "CENTER"
        }
        worksheet.format('A1:G1', {"textFormat": {"bold": True, "fontSize": 14}})
        
        # Add headers in row 2
        headers = ['Item #', 'Exceptions', 'Qty', 'Total Sell', 'Period', 'Exceptions', 'Qty']
        worksheet.update('A2:G2', [headers])
        
        # Format header row
        header_format = {
            "backgroundColor": {"red": 0.83, "green": 0.83, "blue": 0.83},
            "textFormat": {"bold": True},
            "horizontalAlignment": "CENTER"
        }
        worksheet.format('A2:G2', header_format)
        
        # Prepare data rows
        rows = []
        for item in data:
            # Extract period from date if available
            date_str = item.get('date', '')
            period = "P00"  # Default
            if date_str:
                try:
                    if '/' in date_str:
                        month = date_str.split('/')[0]
                        if month.isdigit():
                            period = f"P{month.zfill(2)}"
                except Exception:
                    pass
            
            # Create row with proper format
            qty_raw = str(item.get('quantity', '1')).strip()

            if qty_raw.endswith('-') and qty_raw[:-1].isdigit():
                qty = int('-' + qty_raw[:-1])
            elif qty_raw.isdigit():
                qty = int(qty_raw)
            else:
                   qty = 1
            

            row = [
                item.get('item_number', ''),   # Item #
                
                qty,                           # Qty
                item.get('price', '0.00'),     # Total Sell
                item.get('period', period),    # Period (use provided period or fallback to date-derived)
               
            ]
            rows.append(row)
        
        # Add data rows starting at row 3
        if rows:
            data_range = f"A3:G{2 + len(rows)}"
            worksheet.update(data_range, rows)
            
            # Format the price column (D) as currency
            price_range = f"D3:D{2 + len(rows)}"
            worksheet.format(price_range, {"numberFormat": {"type": "CURRENCY"}})
            
            # Add Grand Total row
            total_row = 3 + len(rows)
            worksheet.update_cell(total_row, 4, "Grand Total")
            worksheet.format(f"D{total_row}", {"textFormat": {"bold": True}})
            
            # Add SUM formula for quantity
            # Note: Google Sheets formulas need to be added with formulas parameter
            qty_sum_formula = f"=SUM(C3:C{total_row-1})"
            worksheet.update_cell(total_row, 7, qty_sum_formula)
        
        # Make the spreadsheet publicly readable
        spreadsheet.share(None, perm_type='anyone', role='reader')
        
        logging.info(f"Data exported to Google Sheets: {spreadsheet.url}")
        return spreadsheet.url
    
    except Exception as e:
        logging.error(f"Error exporting to Google Sheets: {str(e)}")
        raise e
